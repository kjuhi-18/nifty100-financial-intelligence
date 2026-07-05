import sys
import sqlite3
import yaml

from pathlib import Path

import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
# ==========================================================
# Project Root
# ==========================================================

ROOT = Path(__file__).resolve().parents[2]

if str(ROOT) not in sys.path:

    sys.path.append(str(ROOT))


# ==========================================================
# Paths
# ==========================================================

DB_PATH = ROOT / "db" / "nifty100.db"

CONFIG_PATH = ROOT / "config" / "screener_config.yaml"

OUTPUT_DIR = ROOT / "output"

OUTPUT_DIR.mkdir(exist_ok=True)


# ==========================================================
# Screener Engine
# ==========================================================

class ScreenerEngine:

    def __init__(self):

        self.conn = sqlite3.connect(DB_PATH)

        self.cursor = self.conn.cursor()

        self.config = None

        self.data = None

        self.financial_ratios = None

        self.market_cap = None

        self.sectors = None

        self.companies = None

        self.profit_loss = None


# ==========================================================
# Load YAML Configuration
# ==========================================================

    def load_config(self):

        print("\nLoading screener configuration...\n")

        with open(

            CONFIG_PATH,

            "r",

            encoding="utf-8"

        ) as file:

            self.config = yaml.safe_load(file)
            

        print("Configuration loaded.")


# ==========================================================
# Load Database Tables
# ==========================================================

    def load_tables(self):

        print("\nLoading database tables...\n")

        self.financial_ratios = pd.read_sql(

            "SELECT * FROM financial_ratios",

            self.conn

        )

        self.profit_loss = pd.read_sql(

            "SELECT * FROM profitandloss",

            self.conn

        )

        self.companies = pd.read_sql(

            "SELECT * FROM companies",

            self.conn

        )

        self.sectors = pd.read_sql(

            "SELECT * FROM sectors",

            self.conn

        )

        try:

            self.market_cap = pd.read_sql(

                "SELECT * FROM market_cap",

                self.conn

            )

        except Exception:

            self.market_cap = pd.DataFrame()

        print(f"Financial Ratios : {len(self.financial_ratios)}")
        print(f"Profit & Loss    : {len(self.profit_loss)}")
        print(f"Companies        : {len(self.companies)}")
        print(f"Sectors          : {len(self.sectors)}")
        print(f"Market Cap       : {len(self.market_cap)}")
# ==========================================================
# Prepare Master Dataset
# ==========================================================

    def prepare_dataset(self):

        print("\nPreparing screener dataset...\n")

        self.data = self.financial_ratios.merge(

            self.profit_loss[

                [

                    "company_id",

                    "year",

                    "sales",

                    "net_profit",

                    "operating_profit"

                ]

            ],

            on=[

                "company_id",

                "year"

            ],

            how="left"

        )

        self.data = self.data.merge(

            self.sectors[

                [

                    "company_id",

                    "broad_sector"

                ]

            ],

            on="company_id",

            how="left"

        )

        if not self.market_cap.empty:

            columns = self.market_cap.columns.tolist()

            if "market_cap" in columns:

                self.data = self.data.merge(

                    self.market_cap[

                        [

                            "company_id",

                            "market_cap"

                        ]

                    ],

                    on="company_id",

                    how="left"

                )

        print(

            f"Rows    : {len(self.data)}"

        )

        print(

            f"Columns : {len(self.data.columns)}"

        )

        print()

        print(

            self.data.head()

        )

    def get_latest_companies(self):

        df = self.data.copy()

    # Ignore TTM rows
        df = df[df["year"] != "TTM"].copy()

        df["year_num"] = (
        df["year"]
        .astype(str)
        .str.extract(r"(\d{4})")[0]
        .astype(int)
        )

        latest = (
            df.sort_values("year_num")
          .groupby("company_id", as_index=False)
          .tail(1)
          .drop(columns="year_num")
        )

        return latest
# ==========================================================
# Generic Threshold Filter
# ==========================================================

    def apply_filter(

        self,

        dataframe,

        column,

        value,

        operator

    ):

        if column not in dataframe.columns:

            return dataframe

        if operator == ">=":

            dataframe = dataframe[

                dataframe[column] >= value

            ]

        elif operator == "<=":

            dataframe = dataframe[

                dataframe[column] <= value

            ]

        elif operator == ">":

            dataframe = dataframe[

                dataframe[column] > value

            ]

        elif operator == "<":

            dataframe = dataframe[

                dataframe[column] < value

            ]

        elif operator == "==":

            dataframe = dataframe[

                dataframe[column] == value

            ]

        return dataframe


# ==========================================================
# Apply Config Filters
# ==========================================================

    def apply_config_filters(self):

        print("\nApplying configured filters...\n")

        df = self.data.copy()

        filters = self.config["filters"]

        if filters["roe_min"]["enabled"]:

            df = self.apply_filter(

                df,

                "return_on_equity_pct",

                filters["roe_min"]["value"],

                ">="

            )

        if filters["free_cash_flow_min"]["enabled"]:

            df = self.apply_filter(

                df,

                "free_cash_flow_cr",

                filters["free_cash_flow_min"]["value"],

                ">="

            )

        if filters["revenue_cagr_5yr_min"]["enabled"]:

            df = self.apply_filter(

                df,

                "revenue_cagr_5yr",

                filters["revenue_cagr_5yr_min"]["value"],

                ">="

            )

        if filters["pat_cagr_5yr_min"]["enabled"]:

            df = self.apply_filter(

                df,

                "pat_cagr_5yr",

                filters["pat_cagr_5yr_min"]["value"],

                ">="

            )

        if filters["operating_profit_margin_min"]["enabled"]:

            df = self.apply_filter(

                df,

                "operating_profit_margin_pct",

                filters["operating_profit_margin_min"]["value"],

                ">="

            )

        if filters["interest_coverage_min"]["enabled"]:

            df = self.apply_filter(

                df,

                "interest_coverage",

                filters["interest_coverage_min"]["value"],

                ">="

            )

        if filters["asset_turnover_min"]["enabled"]:

            df = self.apply_filter(

                df,

                "asset_turnover",

                filters["asset_turnover_min"]["value"],

                ">="

            )

        if filters["sales_min"]["enabled"]:

            df = self.apply_filter(

                df,

                "sales",

                filters["sales_min"]["value"],

                ">="

            )

        if filters["net_profit_min"]["enabled"]:

            df = self.apply_filter(

                df,

                "net_profit",

                filters["net_profit_min"]["value"],

                ">="

            )

        if filters["eps_cagr_5yr_min"]["enabled"]:

            df = self.apply_filter(

                df,

                "eps_cagr_5yr",

                filters["eps_cagr_5yr_min"]["value"],

                ">="

            )

        self.filtered_data = df

        print(

            f"Filtered Rows : {len(df)}"

        )
# ==========================================================
# Financial Sector D/E Exception
# ==========================================================

    def apply_debt_to_equity_filter(self):

        filters = self.config["filters"]

        if not filters["debt_to_equity_max"]["enabled"]:

            return

        print("\nApplying Debt-to-Equity filter...\n")

        limit = filters["debt_to_equity_max"]["value"]

        financials = self.filtered_data[

            self.filtered_data["broad_sector"] == "Financials"

        ]

        non_financials = self.filtered_data[

            self.filtered_data["broad_sector"] != "Financials"

        ]

        non_financials = non_financials[

            non_financials["debt_to_equity"] <= limit

        ]

        self.filtered_data = pd.concat(

            [

                financials,

                non_financials

            ],

            ignore_index=True

        )

        print(

            f"Rows after D/E filter : {len(self.filtered_data)}"

        )


# ==========================================================
# Interest Coverage Filter
# Debt Free = Infinity
# ==========================================================

    def apply_interest_coverage_logic(self):

        filters = self.config["filters"]

        if not filters["interest_coverage_min"]["enabled"]:

            return

        print("\nApplying Interest Coverage filter...\n")

        minimum = filters["interest_coverage_min"]["value"]

        df = self.filtered_data.copy()

        df["interest_coverage"] = (

            df["interest_coverage"]

            .replace(

                "Debt Free",

                float("inf")

            )

        )

        df["interest_coverage"] = pd.to_numeric(

            df["interest_coverage"],

            errors="coerce"

        )

        df = df[

            df["interest_coverage"] >= minimum

        ]

        self.filtered_data = df

        print(

            f"Rows after ICR filter : {len(df)}"

        )


# ==========================================================
# Optional Filters
# ==========================================================

    def apply_optional_filters(self):

        filters = self.config["filters"]

        df = self.filtered_data.copy()

        optional_filters = [

            ("market_cap", "market_cap_min", ">="),

            ("pe_ratio", "pe_max", "<="),

            ("pb_ratio", "pb_max", "<="),

            ("dividend_yield", "dividend_yield_min", ">=")

        ]

        for column, config_key, operator in optional_filters:

            if column not in df.columns:

                continue

            if not filters[config_key]["enabled"]:

                continue

            df = self.apply_filter(

                df,

                column,

                filters[config_key]["value"],

                operator

            )

        self.filtered_data = df

        print(

            f"Rows after optional filters : {len(df)}"

        )
# ==========================================================
# Reset Filters
# ==========================================================

    def reset_filters(self):

        self.filtered_data = self.data.copy()

# ==========================================================
# Winsorisation (P10/P90)
# ==========================================================

    def winsorize(self, series):

        series = pd.to_numeric(

            series,

            errors="coerce"

        )

        p10 = series.quantile(0.10)

        p90 = series.quantile(0.90)

        return series.clip(

            lower=p10,

            upper=p90

        )


# ==========================================================
# Normalise to 0-100
# ==========================================================

    def normalize_metric(

        self,

        series,

        inverse=False

    ):

        series = self.winsorize(series)

        minimum = series.min()

        maximum = series.max()

        if pd.isna(minimum) or pd.isna(maximum):

            return pd.Series(

                0,

                index=series.index

            )

        if maximum == minimum:

            return pd.Series(

                100,

                index=series.index

            )

        score = (

            (series - minimum)

            /

            (maximum - minimum)

        ) * 100

        if inverse:

            score = 100 - score

        return score.fillna(0)


# ==========================================================
# Metric Score Generator
# ==========================================================

    def generate_metric_scores(self):

        print("\nGenerating normalized metric scores...\n")

        df = self.data.copy()

        metrics = {

            "roe_score":

                ("return_on_equity_pct", False),

            "roce_score":

                ("return_on_capital_employed_pct", False),

            "npm_score":

                ("net_profit_margin_pct", False),

            "fcf_score":

                ("free_cash_flow_cr", False),

            "cfo_score":

                ("cash_from_operations_cr", False),

            "revenue_score":

                ("revenue_cagr_5yr", False),

            "pat_score":

                ("pat_cagr_5yr", False),

            "de_score":

                ("debt_to_equity", True),

            "icr_score":

                ("interest_coverage", False)

        }

        for score_name, (column, inverse) in metrics.items():

            if column in df.columns:

                df[score_name] = self.normalize_metric(

                    df[column],

                    inverse=inverse

                )

            else:

                df[score_name] = 0

        self.data = df

        print(

            df[

                [

                    c for c in df.columns

                    if c.endswith("_score")

                ]

            ].head()

        )
    # ==========================================================
# Composite Quality Score (0-100)
# ==========================================================

    def compute_composite_score(self):

        print("\nComputing Composite Quality Score...\n")

        df = self.data.copy()

        # --------------------------------------------------
        # Profitability (35%)
        # --------------------------------------------------

        profitability = (

            df["roe_score"] * 0.15 +

            df["roce_score"] * 0.10 +

            df["npm_score"] * 0.10

        )

        # --------------------------------------------------
        # Cash Quality (30%)
        # --------------------------------------------------

        fcf_positive = (

            (df["free_cash_flow_cr"] > 0)

            .astype(float)

            * 100

        )

        cfo_pat_ratio = (

            df["cash_from_operations_cr"]

            /

            df["net_profit"]

        )

        cfo_pat_score = self.normalize_metric(

            cfo_pat_ratio

        )

        cash_quality = (

            df["fcf_score"] * 0.15 +

            cfo_pat_score * 0.10 +

            fcf_positive * 0.05

        )

        # --------------------------------------------------
        # Growth (20%)
        # --------------------------------------------------

        growth = (

            df["revenue_score"] * 0.10 +

            df["pat_score"] * 0.10

        )

        # --------------------------------------------------
        # Leverage (15%)
        # --------------------------------------------------

        leverage = (

            df["de_score"] * 0.10 +

            df["icr_score"] * 0.05

        )

        # --------------------------------------------------
        # Final Composite Score
        # --------------------------------------------------

        df["composite_score"] = (

            profitability +

            cash_quality +

            growth +

            leverage

        ).round(2)

        df["composite_score"] = df["composite_score"].clip(

            lower=0,

            upper=100

        )

        self.data = df

        print(

            df[

                [

                    "company_id",

                    "year",

                    "composite_score"

                ]

            ].head()

        )


# ==========================================================
# Composite Score Verification
# ==========================================================

    def verify_composite_score(self):

        print("\n" + "=" * 60)

        print("COMPOSITE SCORE VERIFICATION")

        print("=" * 60)

        print(

            f"Minimum Score : {self.data['composite_score'].min():.2f}"

        )

        print(

            f"Maximum Score : {self.data['composite_score'].max():.2f}"

        )

        print(

            f"Average Score : {self.data['composite_score'].mean():.2f}"

        )

        print("\nTop 10 Composite Scores\n")

        preview = (

            self.data

            .sort_values(

                "composite_score",

                ascending=False

            )

            [

                [

                    "company_id",

                    "year",

                    "composite_score"

                ]

            ]

            .head(10)

        )

        print(

            preview.to_string(index=False)

        )
# ==========================================================
# Sector Relative Composite Score
# ==========================================================

    def compute_sector_relative_score(self):

        print("\nComputing Sector Relative Composite Score...\n")

        df = self.data.copy()

        df["sector_composite_score"] = 0.0

        sectors = sorted(

            df["broad_sector"]

            .dropna()

            .unique()

        )

        for sector in sectors:

            mask = (

                df["broad_sector"] == sector

            )

            sector_scores = self.normalize_metric(

                df.loc[mask, "composite_score"]

            )

            df.loc[mask, "sector_composite_score"] = (

                sector_scores

            )

        self.data = df

        print(

            df[

                [

                    "company_id",

                    "broad_sector",

                    "composite_score",

                    "sector_composite_score"

                ]

            ]

            .head(10)

        )


# ==========================================================
# Sector Leaderboard
# ==========================================================

    def sector_leaderboard(self):

        print("\n" + "=" * 70)

        print("SECTOR LEADERBOARD")

        print("=" * 70)
        latest = self.get_latest_companies()

        leaderboard = (

            latest

            .sort_values(

                [

                    "broad_sector",

                    "sector_composite_score"

                ],

                ascending=[

                    True,

                    False

                ]

            )

            .groupby(

                "broad_sector",

                as_index=False

            )

            .head(3)

        )

        print(

            leaderboard[

                [

                    "broad_sector",

                    "company_id",

                    "sector_composite_score"

                ]

            ]

            .to_string(index=False)

        )


# ==========================================================
# Composite Score Statistics
# ==========================================================

    def composite_statistics(self):

        print("\n" + "=" * 70)

        print("COMPOSITE SCORE STATISTICS")

        print("=" * 70)

        print(

            self.data[

                [

                    "composite_score",

                    "sector_composite_score"

                ]

            ]

            .describe()

        )

        print("\nTop 10 Overall\n")

        print(

            self.data

            .sort_values(

                "composite_score",

                ascending=False

            )

            [

                [

                    "company_id",

                    "composite_score",

                    "sector_composite_score"

                ]

            ]

            .head(10)

            .to_string(index=False)

        )
# ==========================================================
# Apply Preset Filters
# ==========================================================

    def preset_filter(self, filters):

        df = self.get_latest_companies()

        # Extract numeric year
        df["year_num"] = (
        df["year"]
        .astype(str)
        .str.extract(r"(\d{4})")[0]
        .fillna(9999)
        .astype(int)
        )

        df = (
        df.sort_values("year_num")
        .groupby("company_id", as_index=False)
        .tail(1)
        .drop(columns="year_num")
)
        for column, operator, value in filters:

            if column not in df.columns:
                continue

            if operator == ">":
                df = df[df[column] > value]

            elif operator == ">=":
                df = df[df[column] >= value]

            elif operator == "<":
                df = df[df[column] < value]

            elif operator == "<=":
                df = df[df[column] <= value]

            elif operator == "==":
                df = df[df[column] == value]

        self.filtered_data = df

        self.apply_debt_to_equity_filter()

        self.sort_results()

        return self.filtered_data


# ==========================================================
# Save Preset
# ==========================================================

    def save_preset(self, name):

        output = OUTPUT_DIR / f"{name.lower().replace(' ','_')}.csv"

        self.filtered_data.to_csv(

            output,

            index=False

        )

        print(f"{name:<25} {len(self.filtered_data):>4} companies")
# ==========================================================
# Quality Compounder
# ==========================================================

    def quality_compounder(self):

        print("\nRunning Quality Compounder...\n")

        filters = [

            ("return_on_equity_pct", ">=", 15),

            ("debt_to_equity", "<=", 1),

            ("free_cash_flow_cr", ">=", 0),

            ("revenue_cagr_5yr", ">=", 10)

        ]

        self.preset_filter(filters)

        self.save_preset("Quality Compounder")


# ==========================================================
# Growth Accelerator
# ==========================================================

    def growth_accelerator(self):

        print("\nRunning Growth Accelerator...\n")

        filters = [

            ("pat_cagr_5yr", ">=", 20),

            ("revenue_cagr_5yr", ">=", 15),

            ("debt_to_equity", "<=", 2)

        ]

        self.preset_filter(filters)

        self.save_preset("Growth Accelerator")


# ==========================================================
# Run First Two Presets
# ==========================================================

    def run_growth_presets(self):

        print("\n" + "=" * 60)
        print("GROWTH PRESET SCREENERS")
        print("=" * 60)

        self.quality_compounder()

        self.growth_accelerator()

        print("\nGrowth preset screening completed.")
# ==========================================================
# Value Pick
# ==========================================================

    def value_pick(self):

        print("\nRunning Value Pick...\n")

        filters = [

            ("pe_ratio", "<", 20),

            ("pb_ratio", "<", 3),

            ("debt_to_equity", "<", 2),

            ("dividend_yield", ">", 1)

        ]

        self.preset_filter(filters)

        self.save_preset("Value Pick")


# ==========================================================
# Dividend Champion
# ==========================================================

    def dividend_champion(self):

        print("\nRunning Dividend Champion...\n")

        filters = [

            ("return_on_equity_pct", ">=", 12),

            ("dividend_payout_ratio_pct", "<=", 80),

            ("free_cash_flow_cr", ">=", 0)

        ]

        self.preset_filter(filters)

        self.save_preset("Dividend Champion")


# ==========================================================
# Run Value Presets
# ==========================================================

    def run_value_presets(self):

        print("\n" + "=" * 60)
        print("VALUE PRESET SCREENERS")
        print("=" * 60)

        self.value_pick()

        self.dividend_champion()

        print("\nValue preset screening completed.")
# ==========================================================
# Debt-Free Blue Chip
# ==========================================================

    def debt_free_blue_chip(self):

        print("\nRunning Debt-Free Blue Chip...\n")

        filters = [

            ("debt_to_equity", "<=", 0.10),

            ("return_on_equity_pct", ">=", 15),

            ("sales", ">=", 5000)

        ]

        self.preset_filter(filters)

        self.save_preset("Debt Free Blue Chip")


# ==========================================================
# Turnaround Watch
# ==========================================================

    def turnaround_watch(self):

        print("\nRunning Turnaround Watch...\n")

        df = self.get_latest_companies()

        # Revenue CAGR (3 Year)
        if "revenue_cagr_3yr" in df.columns:

            df = df[

                df["revenue_cagr_3yr"] > 10

            ]

        else:

            print(

                "Revenue CAGR 3yr not available - skipping filter"

            )

        # Positive Free Cash Flow
        if "free_cash_flow_cr" in df.columns:

            df = df[

                df["free_cash_flow_cr"] > 0

            ]

        filters = [

        ("free_cash_flow_cr", ">=", 0),

        ("revenue_cagr_5yr", ">=", 8),

        ("return_on_equity_pct", ">=", 10)

]


# ==========================================================
# Run Stability Presets
# ==========================================================

    def run_stability_presets(self):

        print("\n" + "=" * 60)

        print("STABILITY PRESET SCREENERS")

        print("=" * 60)

        self.debt_free_blue_chip()

        self.turnaround_watch()

        print(

            "\nStability preset screening completed."

        )
# ==========================================================
# Run All Presets
# ==========================================================

    def run_all_presets(self):

        print("\n" + "=" * 70)
        print("DAY 16 - RUNNING ALL PRESET SCREENERS")
        print("=" * 70)

        self.preset_results = {}

        preset_functions = [

            ("Quality Compounder", self.quality_compounder),

            ("Value Pick", self.value_pick),

            ("Growth Accelerator", self.growth_accelerator),

            ("Dividend Champion", self.dividend_champion),

            ("Debt Free Blue Chip", self.debt_free_blue_chip),

            ("Turnaround Watch", self.turnaround_watch)

        ]

        for preset_name, preset_function in preset_functions:

            preset_function()

            self.preset_results[preset_name] = len(self.filtered_data)

        print("\nAll preset screeners executed successfully.")


# ==========================================================
# Validate Presets
# ==========================================================

    def validate_presets(self):

        print("\n" + "=" * 70)
        print("PRESET VALIDATION")
        print("=" * 70)

        for preset, count in self.preset_results.items():

            if 5 <= count <= 50:
                status = "PASS"

            elif count == 0:
                status = "CHECK FILTERS"

            elif count < 5:
                status = "LOW"

            else:
                status = "HIGH"

            print(f"{preset:<30} {count:>4} companies   {status}")

        print("\nBusiness sense review completed.")

# ==========================================================
# Export Screener Workbook
# ==========================================================

    def export_screener_workbook(self):

        print("\nGenerating screener_output.xlsx...\n")

        wb = Workbook()

        wb.remove(wb.active)

        green_fill = PatternFill(
            fill_type="solid",
            start_color="C6EFCE",
            end_color="C6EFCE"
        )

        red_fill = PatternFill(
            fill_type="solid",
            start_color="FFC7CE",
            end_color="FFC7CE"
        )

        presets = {

            "Quality Compounder": self.quality_compounder,

            "Value Pick": self.value_pick,

            "Growth Accelerator": self.growth_accelerator,

            "Dividend Champion": self.dividend_champion,

            "Debt Free Blue Chip": self.debt_free_blue_chip,

            "Turnaround Watch": self.turnaround_watch

        }

        columns = [

            "company_id",

            "year",

            "broad_sector",

            "return_on_equity_pct",

            "net_profit_margin_pct",

            "debt_to_equity",

            "interest_coverage",

            "asset_turnover",

            "free_cash_flow_cr",

            "cash_from_operations_cr",

            "revenue_cagr_5yr",

            "pat_cagr_5yr",

            "earnings_per_share",

            "book_value_per_share",

            "sales",

            "net_profit",

            "operating_profit",

            "composite_score",

            "sector_composite_score"

        ]

        for preset_name, preset in presets.items():

            preset()

            ws = wb.create_sheet(

                title=preset_name[:31]

            )

            available = [

                c for c in columns

                if c in self.filtered_data.columns

            ]

            ws.append(available)

            for cell in ws[1]:

                cell.font = Font(bold=True)

            data = (

                self.filtered_data

                .sort_values(

                    "composite_score",

                    ascending=False

                )

            )

            for _, row in data.iterrows():

                values = [

                    row[col]

                    for col in available

                ]

                ws.append(values)

            for row in ws.iter_rows(

                min_row=2

            ):

                for cell in row:

                    if isinstance(

                        cell.value,

                        (int, float)

                    ):

                        if cell.value >= 50:

                            cell.fill = green_fill

                        else:

                            cell.fill = red_fill

        output_file = OUTPUT_DIR / "screener_output.xlsx"

        wb.save(output_file)

        print(

            f"Workbook saved : {output_file}"

        )
# ==========================================================
# Verify Workbook
# ==========================================================

    def verify_screener_output(self):

        print("\n" + "=" * 60)
        print("SCREENER OUTPUT VERIFICATION")
        print("=" * 60)

        output_file = OUTPUT_DIR / "screener_output.xlsx"

        if output_file.exists():

            print("Workbook              PASS")
            print(f"Location : {output_file}")

        else:

            print("Workbook              FAIL")

        print()

        print("Composite Score Range")

        print("------------------------------")

        print(
            f"Minimum : {self.data['composite_score'].min():.2f}"
        )

        print(
            f"Maximum : {self.data['composite_score'].max():.2f}"
        )

        print(
            f"Average : {self.data['composite_score'].mean():.2f}"
        )

        print()

        print("Sector Relative Score")

        print("------------------------------")

        print(
            f"Minimum : {self.data['sector_composite_score'].min():.2f}"
        )

        print(
            f"Maximum : {self.data['sector_composite_score'].max():.2f}"
        )

        print(
            f"Average : {self.data['sector_composite_score'].mean():.2f}"
        )


# ==========================================================
# Day 17 Summary
# ==========================================================

    def day17_summary(self):

        print("\n" + "=" * 70)
        print("SPRINT 3 - DAY 17 COMPLETED")
        print("=" * 70)

        print("\nDeliverables")
        print("----------------------------------------")

        print("✓ Winsorisation implemented")
        print("✓ Metric normalisation completed")
        print("✓ Composite quality score (0-100)")
        print("✓ Sector-relative score")
        print("✓ Screener workbook generated")
        print("✓ Six preset worksheets exported")
        print("✓ Composite score verification completed")

        print("\nOutput Files")
        print("----------------------------------------")

        print("✓ screener_output.xlsx")

        print("\nReady for Day 18")

# ==========================================================
# Final Sorting
# ==========================================================

    def sort_results(self):

        print("\nSorting results...\n")

        if "composite_quality_score" in self.filtered_data.columns:

            self.filtered_data = self.filtered_data.sort_values(

                "composite_quality_score",

                ascending=False

            )

        elif "return_on_equity_pct" in self.filtered_data.columns:

            self.filtered_data = self.filtered_data.sort_values(

                "return_on_equity_pct",

                ascending=False

            )

        self.filtered_data = self.filtered_data.reset_index(

            drop=True

        )

        print(

            self.filtered_data.head(10)

        )
# ==========================================================
# Screener Summary
# ==========================================================

    def screener_summary(self):

        print("\n" + "=" * 60)
        print("SCREENER SUMMARY")
        print("=" * 60)

        print(f"Companies Selected : {len(self.filtered_data)}")

        if len(self.filtered_data) == 0:

            print("\nNo companies matched the filters.")
            return

        latest = self.filtered_data.copy()

        if "year" in latest.columns:

            latest = latest.sort_values(
                ["company_id", "year"]
            )

        print("\nTop 10 Companies\n")

        preview_columns = [

            "company_id",
            "year",
            "return_on_equity_pct",
            "debt_to_equity",
            "free_cash_flow_cr",
            "revenue_cagr_5yr",
            "pat_cagr_5yr",
            "composite_quality_score"

        ]

        preview_columns = [

            c for c in preview_columns
            if c in latest.columns

        ]

        print(

            latest[preview_columns]

            .head(10)

            .to_string(index=False)

        )


# ==========================================================
# Export Results
# ==========================================================

    def export_results(self):

        print("\nExporting screener results...\n")

        output_file = OUTPUT_DIR / "screener_results.csv"

        self.filtered_data.to_csv(

            output_file,

            index=False

        )

        print(

            f"Saved : {output_file}"

        )


# ==========================================================
# Verification
# ==========================================================

    def verify_results(self):

        print("\n" + "=" * 60)
        print("VERIFICATION")
        print("=" * 60)

        print(

            f"Rows Returned : {len(self.filtered_data)}"

        )

        print(

            f"Columns : {len(self.filtered_data.columns)}"

        )

        required = [

            "company_id",
            "year",
            "return_on_equity_pct",
            "debt_to_equity",
            "free_cash_flow_cr",
            "revenue_cagr_5yr",
            "pat_cagr_5yr",
            "composite_quality_score"

        ]

        print("\nChecking required columns...\n")

        for column in required:

            if column in self.filtered_data.columns:

                print(f"{column:<35} PASS")

            else:

                print(f"{column:<35} MISSING")

        print("\nChecking duplicate rows...\n")

        duplicates = self.filtered_data.duplicated(

            subset=[

                "company_id",

                "year"

            ]

        ).sum()

        print(

            f"Duplicate Company-Year Rows : {duplicates}"

        )

        print("\nChecking NULL counts...\n")

        nulls = self.filtered_data[required].isna().sum()

        print(nulls)

        print("\nVerification Complete.")
# ==========================================================
# Close Connection
# ==========================================================

    def close(self):

        if self.conn:

            self.conn.close()


# ==========================================================
# Main
# ==========================================================

def main():

    engine = ScreenerEngine()

    try:

        engine.load_config()

        engine.load_tables()

        engine.prepare_dataset()

        engine.apply_config_filters()

        engine.apply_debt_to_equity_filter()

        engine.apply_interest_coverage_logic()

        engine.apply_optional_filters()

        engine.sort_results()

        engine.screener_summary()

        engine.export_results()

        engine.verify_results()

        engine.run_all_presets()

        engine.validate_presets()
        engine.generate_metric_scores()

        engine.compute_composite_score()

        engine.verify_composite_score()

        engine.compute_sector_relative_score()

        engine.sector_leaderboard()

        engine.composite_statistics()

        engine.export_screener_workbook()

        engine.verify_screener_output()

        engine.day17_summary()





    except Exception as e:

        print("\nERROR")
        print("-" * 40)

        print(e)

        raise

    finally:

        engine.close()


# ==========================================================
# Entry Point
# ==========================================================

if __name__ == "__main__":

    main()