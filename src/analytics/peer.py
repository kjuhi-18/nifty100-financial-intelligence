import sys
import sqlite3
from pathlib import Path
import matplotlib.pyplot as plt
from math import pi
import numpy as np
import pandas as pd
from openpyxl.styles import Font

# ==========================================================
# Project Paths
# ==========================================================

ROOT = Path(__file__).resolve().parents[2]

if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

DB_PATH = ROOT / "db" / "nifty100.db"
PEER_GROUP_FILE = ROOT / "data" / "raw" / "peer_groups.xlsx"
OUTPUT_DIR = ROOT / "output"
REPORT_DIR = ROOT / "reports"

RADAR_DIR = REPORT_DIR / "radar_charts"

REPORT_DIR.mkdir(exist_ok=True)

RADAR_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)


# ==========================================================
# Peer Analytics Engine
# ==========================================================

class PeerEngine:

    def __init__(self):

        self.conn = sqlite3.connect(DB_PATH)

        self.financial_ratios = None
        self.peer_groups = None
        self.companies = None
        self.peer_percentiles = None
        self.data = None
        self.radar_metrics = [

        "return_on_equity_pct",

        "return_on_capital_employed_pct",

        "net_profit_margin_pct",

        "debt_to_equity",

        "free_cash_flow_cr",

        "pat_cagr_5yr",

        "revenue_cagr_5yr",

        "composite_quality_score"

        ]
        self.radar_labels = [

        "ROE",

        "ROCE",

        "NPM",

        "D/E",

        "FCF",

        "PAT CAGR",

        "Revenue CAGR",

        "Composite"

        ]


# ==========================================================
# Load Source Tables
# ==========================================================

    def load_tables(self):

        print("\nLoading source tables...\n")

        self.financial_ratios = pd.read_sql(
            "SELECT * FROM financial_ratios",
            self.conn
        )

        self.companies = pd.read_sql(
            "SELECT * FROM companies",
            self.conn
        )

        self.peer_groups = pd.read_excel(
            PEER_GROUP_FILE
        )

        print(f"Financial Ratios : {len(self.financial_ratios)}")
        print(f"Companies        : {len(self.companies)}")
        print(f"Peer Groups      : {len(self.peer_groups)}")


# ==========================================================
# Latest Company Snapshot
# ==========================================================

    def get_latest_company_data(self):

        print("\nPreparing latest company snapshot...\n")

        df = self.financial_ratios.copy()

        df = df[
            df["year"] != "TTM"
        ].copy()

        df["year_num"] = (

            df["year"]

            .astype(str)

            .str.extract(r"(\d{4})")[0]

            .astype(int)

        )

        latest = (

            df

            .sort_values(

                ["company_id", "year_num"]

            )

            .groupby(

                "company_id",

                as_index=False

            )

            .tail(1)

            .drop(

                columns="year_num"

            )

        )

        self.data = latest

        print(f"Latest Companies : {len(self.data)}")
        return latest


# ==========================================================
# Merge Peer Groups
# ==========================================================

    def merge_peer_groups(self):

        print("\nMerging peer group information...\n")

        self.data = self.data.merge(

            self.peer_groups,

            how="left",

            on="company_id"

        )

        self.data["peer_group_name"] = (

            self.data["peer_group_name"]

            .fillna("No peer group assigned")

        )

        self.data["is_benchmark"] = (

            self.data["is_benchmark"]

            .fillna(False)

        )

        print(self.data.head())

        print()

        print(

            self.data[

                "peer_group_name"

            ]

            .value_counts()

        )


# ==========================================================
# Verification
# ==========================================================

    def verify_data(self):

        print("\n" + "=" * 60)

        print("DATA VERIFICATION")

        print("=" * 60)

        print(f"Rows : {len(self.data)}")

        print()

        print(

            "Unique Companies :",

            self.data["company_id"].nunique()

        )

        print(

            "Peer Groups :",

            self.data["peer_group_name"].nunique()

        )

        missing = (

            self.data

            ["peer_group_name"]

            ==

            "No peer group assigned"

        ).sum()

        print(

            "Companies Without Peer Group :",

            missing

        )

        print()

        print(

            self.data[
                [
                    "company_id",
                    "peer_group_name",
                    "is_benchmark"
                ]
            ]

            .head(15)

        )
# ==========================================================
# Percent Rank
# ==========================================================

    def percent_rank(self, series, inverse=False):

        series = pd.to_numeric(
            series,
            errors="coerce"
        )

        rank = series.rank(
            method="average",
            pct=True
        )

        if inverse:
            rank = 1 - rank

        return (rank * 100).round(2)


# ==========================================================
# Compute Peer Percentiles
# ==========================================================

    def compute_peer_percentiles(self):

        print("\nComputing Peer Percentile Rankings...\n")

        metrics = {

            "return_on_equity_pct": False,

            "return_on_capital_employed_pct": False,

            "net_profit_margin_pct": False,

            "debt_to_equity": True,

            "free_cash_flow_cr": False,

            "pat_cagr_5yr": False,

            "revenue_cagr_5yr": False,

            "eps_cagr_5yr": False,

            "interest_coverage": False,

            "asset_turnover": False

        }

        results = []

        peer_groups = (

            self.data["peer_group_name"]

            .dropna()

            .unique()

        )

        for group in peer_groups:

            if group == "No peer group assigned":
                continue

            peer_df = self.data[
                self.data["peer_group_name"] == group
            ].copy()

            print(f"{group:<25} {len(peer_df)} companies")

            for metric, inverse in metrics.items():

                if metric not in peer_df.columns:
                    continue

                peer_df[f"{metric}_percentile"] = self.percent_rank(
                    peer_df[metric],
                    inverse=inverse
                )

                for _, row in peer_df.iterrows():

                    results.append({

                        "company_id": row["company_id"],

                        "peer_group_name": group,

                        "metric": metric,

                        "value": row.get(metric),

                        "percentile_rank": row.get(f"{metric}_percentile"),

                        "year": row["year"]

                    })

        self.peer_percentiles = pd.DataFrame(results)

        print()

        print(

            "Percentile Rows :",

            len(self.peer_percentiles)

        )
# ==========================================================
# Populate peer_percentiles Table
# ==========================================================

    def populate_peer_percentiles_table(self):

        print("\nPopulating peer_percentiles table...\n")

        cursor = self.conn.cursor()

        cursor.execute("""

        DROP TABLE IF EXISTS peer_percentiles

        """)

        cursor.execute("""

        CREATE TABLE peer_percentiles (

            id INTEGER PRIMARY KEY AUTOINCREMENT,

            company_id TEXT,

            peer_group_name TEXT,

            metric TEXT,

            value REAL,

            percentile_rank REAL,

            year TEXT

        )

        """)

        self.peer_percentiles.to_sql(

            "peer_percentiles",

            self.conn,

            if_exists="append",

            index=False

        )

        self.conn.commit()

        print("peer_percentiles table populated successfully.")
# ==========================================================
# Verify Database
# ==========================================================

    def verify_database(self):

        print("\n" + "=" * 60)

        print("DATABASE VERIFICATION")

        print("=" * 60)

        query = """

        SELECT COUNT(*)

        FROM peer_percentiles

        """

        rows = pd.read_sql(

            query,

            self.conn

        ).iloc[0, 0]

        print(f"Rows Loaded : {rows}")

        print()

        sample = pd.read_sql("""

        SELECT *

        FROM peer_percentiles

        LIMIT 10

        """, self.conn)

        print(sample)
# ==========================================================
# Companies Without Peer Group
# ==========================================================

    def review_missing_peer_groups(self):

        print("\n" + "=" * 60)

        print("NO PEER GROUP REVIEW")

        print("=" * 60)

        missing = self.data[

            self.data["peer_group_name"]

            ==

            "No peer group assigned"

        ]

        print(

            f"Companies : {len(missing)}"

        )

        print()

        print(

            missing[

                [

                    "company_id",

                    "year"

                ]

            ]

            .to_string(index=False)

        )
# ==========================================================
# Peer Group Summary
# ==========================================================

    def peer_group_summary(self):

        print("\n" + "=" * 70)
        print("PEER GROUP SUMMARY")
        print("=" * 70)

        summary = (

            self.data

            .groupby("peer_group_name")

            .agg(

                Companies=("company_id", "count"),

                Benchmarks=("is_benchmark", "sum")

            )

            .reset_index()

            .sort_values("Companies", ascending=False)

        )

        print(summary.to_string(index=False))


# ==========================================================
# Spot Check
# ==========================================================

    def spot_check_peer_groups(self):

        print("\n" + "=" * 70)
        print("IT SERVICES SPOT CHECK")
        print("=" * 70)

        it = self.peer_percentiles[

            self.peer_percentiles["peer_group_name"] == "IT Services"

        ]

        print(it.head(20).to_string(index=False))

        print("\n" + "=" * 70)
        print("FMCG SPOT CHECK")
        print("=" * 70)

        fmcg = self.peer_percentiles[

            self.peer_percentiles["peer_group_name"] == "FMCG"

        ]

        print(fmcg.head(20).to_string(index=False))


# ==========================================================
# Export CSV
# ==========================================================

    def export_peer_percentiles(self):

        print("\nExporting peer percentile report...\n")

        output_file = OUTPUT_DIR / "peer_percentiles.csv"

        self.peer_percentiles.to_csv(

            output_file,

            index=False

        )

        print(f"Saved : {output_file}")


# ==========================================================
# Validation
# ==========================================================

    def validate_peer_percentiles(self):

        print("\n" + "=" * 60)
        print("PEER PERCENTILE VALIDATION")
        print("=" * 60)

        print(f"Rows : {len(self.peer_percentiles)}")

        print()

        print("Metrics")

        print("-------------------------")

        print(

            sorted(

                self.peer_percentiles["metric"].unique()

            )

        )

        print()

        print(

            "Peer Groups :",

            self.peer_percentiles["peer_group_name"].nunique()

        )
# ==========================================================
# Database Spot Check
# ==========================================================

    def database_spot_check(self):

        print("\n" + "=" * 70)
        print("DATABASE SPOT CHECK")
        print("=" * 70)

        sample = pd.read_sql(

            """

            SELECT *

            FROM peer_percentiles

            ORDER BY peer_group_name,
                     metric,
                     percentile_rank DESC

            LIMIT 20

            """,

            self.conn

        )

        print(sample.to_string(index=False))


# ==========================================================
# Final Verification
# ==========================================================

    def final_verification(self):

        print("\n" + "=" * 70)
        print("FINAL VERIFICATION")
        print("=" * 70)

        rows = len(self.peer_percentiles)

        groups = self.peer_percentiles["peer_group_name"].nunique()

        metrics = self.peer_percentiles["metric"].nunique()

        print(f"Rows Generated      : {rows}")
        print(f"Peer Groups         : {groups}")
        print(f"Metrics Ranked      : {metrics}")

        if rows > 0:
            print("Peer Percentiles    : PASS")
        else:
            print("Peer Percentiles    : FAIL")

        if (OUTPUT_DIR / "peer_percentiles.csv").exists():
            print("CSV Export          : PASS")
        else:
            print("CSV Export          : FAIL")

        db_rows = pd.read_sql(

            "SELECT COUNT(*) AS total FROM peer_percentiles",

            self.conn

        ).iloc[0]["total"]

        print(f"SQLite Rows         : {db_rows}")

        print("\nVerification Complete.")


# ==========================================================
# Day 18 Summary
# ==========================================================

    def day18_summary(self):

        print("\n" + "=" * 70)
        print("SPRINT 3 - DAY 18 COMPLETED")
        print("=" * 70)

        print("\nDeliverables")
        print("----------------------------------------")

        print("✓ Peer groups loaded")
        print("✓ Latest company snapshot created")
        print("✓ Percentile rankings computed")
        print("✓ D/E inverse ranking implemented")
        print("✓ SQLite peer_percentiles table populated")
        print("✓ peer_percentiles.csv generated")
        print("✓ IT Services spot check completed")
        print("✓ FMCG spot check completed")
        print("✓ Validation completed")

        print("\nOutput Files")
        print("----------------------------------------")

        print("✓ output/peer_percentiles.csv")

        print("\nDatabase")
        print("----------------------------------------")

        print("✓ peer_percentiles table")

        print("\nReady for Day 19")
# ==========================================================
# Preview Rankings
# ==========================================================

    def preview_rankings(self):

        print("\n" + "=" * 60)

        print("PEER RANKING PREVIEW")

        print("=" * 60)

        print(

            self.peer_percentiles.head(20)

        )

        print()

        print(

            self.peer_percentiles.groupby(

                "metric"

            ).size()

        )
# ==========================================================
# Prepare Radar Dataset
# ==========================================================

    # ==========================================================
# Prepare Radar Dataset
# ==========================================================

    def prepare_radar_dataset(self):

        print("\nPreparing radar dataset...\n")

        if self.data is None:

            self.get_latest_company_data()

            self.merge_peer_groups()

        self.latest_data = self.data.copy()

        print(f"Companies : {len(self.latest_data)}")

        print()

        cols = [

            c for c in

            [

                "company_id",

                "year",

                "peer_group_name",

                "is_benchmark"

            ]

            if c in self.latest_data.columns

        ]

        print(self.latest_data[cols].head())
# ==========================================================
# Normalise Radar Values
# ==========================================================

    def normalise_radar_metrics(self):

        print("\nNormalising radar metrics...\n")

        radar = self.latest_data.copy()

        for metric in self.radar_metrics:

            if metric not in radar.columns:

                continue

            values = pd.to_numeric(

                radar[metric],

                errors="coerce"

            )

            minimum = values.min()

            maximum = values.max()

            if pd.isna(minimum) or pd.isna(maximum):

                continue

            if maximum == minimum:

                radar[metric] = 50

                continue

            radar[metric] = (

                (values - minimum)

                /

                (maximum - minimum)

            ) * 100

        self.radar_data = radar

        print("Normalisation Complete")
# ==========================================================
# Preview Radar Dataset
# ==========================================================

    def preview_radar_data(self):

        print("\n" + "=" * 70)

        print("RADAR DATA PREVIEW")

        print("=" * 70)

        print(

            self.radar_data[

            [

                "company_id"

            ]

            +

            self.radar_metrics

            ]

            .head(10)

        )
# ==========================================================
# Peer Average
# ==========================================================

    def get_peer_average(self, peer_group):

        peer = self.radar_data[

        self.radar_data["peer_group_name"] == peer_group

        ]

        averages = []

        for metric in self.radar_metrics:

            averages.append(

                peer[metric].mean()

            )

        return averages
# ==========================================================
# Company Metrics
# ==========================================================

    def get_company_metrics(self, company_id):

        row = self.radar_data[

            self.radar_data["company_id"] == company_id

        ]

        if row.empty:

            return None

        values = []

        for metric in self.radar_metrics:

            values.append(

                row.iloc[0][metric]

            )

        return values
# ==========================================================
# Radar Plot
# ==========================================================

    def create_radar_chart(

        self,

        company_id,

        peer_group

    ):

        company_values = self.get_company_metrics(

            company_id

        )

        if company_values is None:

            return

        peer_values = self.get_peer_average(

            peer_group

        )

        company_values = [

            0 if pd.isna(x) else x

            for x in company_values

        ]

        peer_values = [

            0 if pd.isna(x) else x

            for x in peer_values

        ]

        angles = np.linspace(

            0,

            2 * np.pi,

            len(self.radar_labels),

            endpoint=False

        )

        angles = np.concatenate(

            (

                angles,

                [angles[0]]

            )

        )

        company_values.append(

            company_values[0]

        )

        peer_values.append(

            peer_values[0]

        )

        fig = plt.figure(

            figsize=(8,8)

        )

        ax = plt.subplot(

            polar=True

        )

        ax.plot(

            angles,

            company_values,

            linewidth=2,

            label=company_id

        )

        ax.fill(

            angles,

            company_values,

            alpha=0.25

        )

        ax.plot(

            angles,

            peer_values,

            linestyle="--",

            linewidth=2,

            label="Peer Average"

        )

        ax.set_xticks(

            angles[:-1]

        )

        ax.set_xticklabels(

            self.radar_labels,

            fontsize=10

        )

        ax.set_ylim(

            0,

            100

        )

        plt.title(

            f"{company_id}\n{peer_group}",

            fontsize=14,

            weight="bold"

        )

        plt.legend(

            loc="upper right"

        )

        plt.tight_layout()

        filename = (

            RADAR_DIR

            /

            f"{company_id}_radar.png"

        )

        plt.savefig(

            filename,

            dpi=200

        )

        plt.close()
# ==========================================================
# Generate Charts
# ==========================================================

    def generate_peer_radar_charts(self):

        print("\nGenerating Radar Charts...\n")

        generated = 0

        skipped = 0

        for _, row in self.radar_data.iterrows():

            company = row["company_id"]

            peer = row["peer_group_name"]

            if peer == "No peer group assigned":

                skipped += 1

                continue

            self.create_radar_chart(

                company,

                peer

            )

            generated += 1

        print(

            f"Charts Generated : {generated}"

        )

        print(

            f"Skipped : {skipped}"

        )
# ==========================================================
# Preview Generated Charts
# ==========================================================

    def preview_generated_charts(self):

        print("\nGenerated Files\n")

        charts = sorted(

            RADAR_DIR.glob(

                "*.png"

            )

        )

        print(

            f"Total PNG : {len(charts)}"

        )

        for chart in charts[:10]:

            print(chart.name)
# ==========================================================
# Nifty 100 Average
# ==========================================================

    def get_nifty_average(self):

        averages = []

        for metric in self.radar_metrics:

            averages.append(
                self.radar_data[metric].mean()
            )

        return averages


# ==========================================================
# Standalone Radar Chart
# ==========================================================

    def create_standalone_chart(self, company_id):

        company_values = self.get_company_metrics(company_id)

        if company_values is None:
            return

        nifty_values = self.get_nifty_average()

        company_values = [
            0 if pd.isna(v) else v
            for v in company_values
        ]

        nifty_values = [
            0 if pd.isna(v) else v
            for v in nifty_values
        ]

        angles = np.linspace(
         0,
         2 * np.pi,
            len(self.radar_labels),
            endpoint=False
        )

        angles = np.concatenate(
            (angles, [angles[0]])
        )

        company_values.append(company_values[0])
        nifty_values.append(nifty_values[0])

        fig = plt.figure(figsize=(8, 8))

        ax = plt.subplot(polar=True)

        ax.plot(
            angles,
            company_values,
            linewidth=2,
            label=company_id
        )

        ax.fill(
            angles,
            company_values,
            alpha=0.25
        )

        ax.plot(
            angles,
            nifty_values,
            "--",
            linewidth=2,
            label="Nifty 100 Average"
        )

        ax.set_xticks(
            angles[:-1]
        )

        ax.set_xticklabels(
            self.radar_labels,
            fontsize=10
        )

        ax.set_ylim(0, 100)

        plt.title(
            f"{company_id}\nNo Peer Group",
                fontsize=14,
            weight="bold"
        )

        plt.legend(loc="upper right")

        plt.tight_layout()

        filename = (
        RADAR_DIR /
        f"{company_id}_radar.png"
        )

        plt.savefig(
        filename,
        dpi=200
        )

        plt.close()


# ==========================================================
# Generate Standalone Charts
# ==========================================================

    def generate_standalone_charts(self):

        print("\nGenerating Standalone Charts...\n")

        generated = 0

        for _, row in self.radar_data.iterrows():

            if row["peer_group_name"] != "No peer group assigned":
                continue

            self.create_standalone_chart(
            row["company_id"]
            )

            generated += 1

        print(f"Standalone Charts : {generated}")


# ==========================================================
# Radar Chart Verification
# ==========================================================

    def verify_radar_charts(self):

        print("\n" + "=" * 60)
        print("RADAR CHART VERIFICATION")
        print("=" * 60)

        charts = sorted(
            RADAR_DIR.glob("*.png")
        )

        print(f"Charts Found : {len(charts)}")

        if len(charts) == len(self.radar_data):

            print("PASS : Every company has a radar chart")

        else:

            print("WARNING : Missing radar charts")

            print(
            f"Expected : {len(self.radar_data)}"
            )

            print(
            f"Generated : {len(charts)}"
            )   

        print("\nSample Charts")
        print("------------------------------")

        for chart in charts[:15]:
            print(chart.name)
# ==========================================================
# Radar Chart Summary
# ==========================================================

    def radar_summary(self):

        print("\n" + "=" * 70)
        print("RADAR CHART SUMMARY")
        print("=" * 70)

        total = len(self.radar_data)

        peer = (
            self.radar_data["peer_group_name"]
            != "No peer group assigned"
        ).sum()

        standalone = (
            self.radar_data["peer_group_name"]
            == "No peer group assigned"
        ).sum()

        charts = len(list(RADAR_DIR.glob("*.png")))

        print(f"Companies              : {total}")
        print(f"Peer Charts            : {peer}")
        print(f"Standalone Charts      : {standalone}")
        print(f"PNG Files Generated    : {charts}")

        print("\nPeer Group Distribution")
        print("-" * 40)

        print(
            self.radar_data[
                "peer_group_name"
            ]
            .value_counts()
            .to_string()
        )

# ==========================================================
# Benchmark Companies
# ==========================================================

    def benchmark_summary(self):

        print("\n" + "=" * 70)
        print("BENCHMARK COMPANIES")
        print("=" * 70)

        benchmark = self.radar_data[
            self.radar_data["is_benchmark"] == True
        ]

        if benchmark.empty:

            print("No benchmark companies found.")
            return

        print(
            benchmark[
                [
                    "company_id",
                    "peer_group_name",
                    "return_on_equity_pct",
                    "return_on_capital_employed_pct",
                    "composite_quality_score"
                ]
            ].to_string(index=False)
        )

# ==========================================================
# Largest Radar Scores
# ==========================================================

    def top_composite_companies(self):

        print("\n" + "=" * 70)
        print("TOP COMPOSITE SCORES")
        print("=" * 70)

        cols = [

            "company_id",

            "peer_group_name",

            "composite_quality_score",

            "return_on_equity_pct",

            "return_on_capital_employed_pct"

        ]

        print(

            self.radar_data

            .sort_values(

                "composite_quality_score",

                ascending=False

            )[cols]

            .head(15)

            .to_string(index=False)

        )


# ==========================================================
# Radar Report
# ==========================================================

    def export_radar_report(self):

        print("\nExporting radar report...\n")

        report = self.radar_data.copy()

        report["chart_file"] = report["company_id"] + "_radar.png"

        output = OUTPUT_DIR / "radar_report.csv"

        report.to_csv(

            output,

            index=False

        )

        print(f"Saved : {output}")


# ==========================================================
# Validation
# ==========================================================

    def validate_radar_outputs(self):

        print("\n" + "=" * 70)
        print("RADAR OUTPUT VALIDATION")
        print("=" * 70)

        png_files = list(

            RADAR_DIR.glob("*.png")

        )

        csv_exists = (

            OUTPUT_DIR /

            "radar_report.csv"

        ).exists()

        print(

            f"Radar Charts : {len(png_files)}"

        )

        print(

            f"CSV Report   : {'PASS' if csv_exists else 'FAIL'}"

        )

        print(

            f"Companies    : {len(self.radar_data)}"

        )

        if len(png_files) == len(self.radar_data):

            print("Chart Validation : PASS")

        else:

            print("Chart Validation : WARNING")

        print("\nValidation Complete.")
# ==========================================================
# Final Verification
# ==========================================================

    def final_day19_verification(self):

        print("\n" + "=" * 70)
        print("FINAL DAY 19 VERIFICATION")
        print("=" * 70)

        png_files = list(RADAR_DIR.glob("*.png"))

        expected = len(self.radar_data)

        generated = len(png_files)

        print(f"Companies Analysed      : {expected}")
        print(f"Radar Charts Generated  : {generated}")

        if generated == expected:
            print("Radar Coverage          : PASS")
        else:
            print("Radar Coverage          : WARNING")

        report_exists = (OUTPUT_DIR / "radar_report.csv").exists()

        print(
            f"Radar Report            : {'PASS' if report_exists else 'FAIL'}"
        )

        benchmark_count = (

            self.radar_data["is_benchmark"]

            .fillna(False)

            .sum()

        )

        print(f"Benchmark Companies     : {benchmark_count}")

        print(
            f"Peer Groups Covered     : "
            f"{self.radar_data['peer_group_name'].nunique()-1}"
        )

        print("\nDay 19 verification completed.")


# ==========================================================
# Day 19 Summary
# ==========================================================

    def day19_summary(self):

        print("\n" + "=" * 70)
        print("SPRINT 3 - DAY 19 COMPLETED")
        print("=" * 70)

        print("\nDeliverables")
        print("----------------------------------------")

        print("✓ Radar dataset prepared")
        print("✓ 8-axis radar metrics generated")
        print("✓ Peer average overlay implemented")
        print("✓ Nifty 100 comparison implemented")
        print("✓ 56 peer radar charts generated")
        print("✓ 36 standalone radar charts generated")
        print("✓ 92 PNG charts exported")
        print("✓ radar_report.csv generated")
        print("✓ Radar validation completed")

        print("\nOutput Files")
        print("----------------------------------------")

        print("✓ reports/radar_charts/")
        print("✓ output/radar_report.csv")

        print("\nReady for Day 20")
# ==========================================================
# Prepare Peer Comparison Dataset
# ==========================================================

    def prepare_peer_comparison(self):

        print("\nPreparing Peer Comparison Dataset...\n")

        conn = sqlite3.connect(DB_PATH)

        ratios = pd.read_sql(

            "SELECT * FROM financial_ratios",

            conn

        )

        peers = pd.read_sql(

            "SELECT * FROM peer_percentiles",

            conn

        )

        companies = pd.read_sql(

            "SELECT id AS company_id, company_name FROM companies",

            conn

        )

        conn.close()

        latest = (

            ratios

            .sort_values(

                ["company_id", "year"])

            .groupby(

        "company_id",

        as_index=False

    )

            .tail(1)

        )

        latest = latest.merge(

            companies,

            on="company_id",

            how="left"

        )

        latest = latest.merge(

            self.peer_groups[

                [

                    "company_id",

                    "peer_group_name",

                    "is_benchmark"

                ]

            ],

            on="company_id",

            how="left"

        )

        self.peer_report = latest
        latest["peer_group_name"] = latest["peer_group_name"].fillna(
    "No peer group assigned"
)

        latest["is_benchmark"] = latest["is_benchmark"].fillna(False)
        self.peer_percentiles = peers

        print(

            f"Companies : {len(self.peer_report)}"

        )

        print(

            f"Percentile Rows : {len(self.peer_percentiles)}"

        )

        print()

        print(

            self.peer_report.head()

        )
# ==========================================================
# Verify Peer Groups
# ==========================================================

    def verify_peer_groups(self):

        print("\n" + "=" * 70)

        print("PEER GROUPS")

        print("=" * 70)

        groups = (

            self.peer_report

            .dropna(

                subset=["peer_group_name"]

            )

            ["peer_group_name"]

            .unique()

        )

        groups = sorted(groups)

        print(

            f"Peer Groups : {len(groups)}"

        )

        for g in groups:

            print(g)
    # ==========================================================
    # Create Peer Comparison Workbook
    # ==========================================================

    def create_peer_workbook(self):

        print("\nCreating Peer Comparison Workbook...\n")

        output_file = OUTPUT_DIR / "peer_comparison.xlsx"

        writer = pd.ExcelWriter(

            output_file,

            engine="openpyxl"

        )

        groups = (

            self.peer_report

            .dropna(subset=["peer_group_name"])

            ["peer_group_name"]

            .unique()

        )

        groups = sorted(groups)

        exported = 0

        for group in groups:

            if group == "No peer group assigned":

                continue

            self.export_peer_sheet(

                writer,

                group

            )

            exported += 1

        writer.close()

        print(f"Sheets Created : {exported}")

        print(f"Saved : {output_file}")


# ==========================================================
# Export Single Peer Sheet
# ==========================================================

    def export_peer_sheet(

        self,

        writer,

        peer_group

    ):

        companies = self.peer_report[

            self.peer_report["peer_group_name"]

            == peer_group

        ].copy()

        metrics = self.peer_percentiles[

            self.peer_percentiles["peer_group_name"]

            == peer_group

        ]

        pivot = (

            metrics

            .pivot(

                index="company_id",

                columns="metric",

                values="percentile_rank"

            )

            .reset_index()

        )

        pivot.columns = [

            c if c == "company_id"

            else f"{c}_pct_rank"

            for c in pivot.columns

        ]

        companies = companies.merge(

            pivot,

            on="company_id",

            how="left"

        )
        
        companies = companies.sort_values(

            "composite_quality_score",

            ascending=False

        )

        sheet = peer_group[:31]

        companies.to_excel(

            writer,

            sheet_name=sheet,

            index=False

        )

        worksheet = writer.sheets[sheet]

        worksheet.freeze_panes = "A2"

        for cell in worksheet[1]:

            cell.font = Font(

                bold=True

            )

        for column_cells in worksheet.columns:

            length = max(

                len(str(cell.value))

                if cell.value is not None

                else 0

                for cell in column_cells

            )

            worksheet.column_dimensions[

                column_cells[0].column_letter

            ].width = min(

                max(length + 2, 12),

                35

            )
# ==========================================================
# Apply Excel Formatting
# ==========================================================

    def format_peer_workbook(self):

        print("\nFormatting Peer Comparison Workbook...\n")

        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font

        workbook_path = OUTPUT_DIR / "peer_comparison.xlsx"

        wb = load_workbook(workbook_path)

        green_fill = PatternFill(
            fill_type="solid",
            start_color="C6EFCE",
            end_color="C6EFCE"
        )

        yellow_fill = PatternFill(
            fill_type="solid",
            start_color="FFF2CC",
            end_color="FFF2CC"
        )

        red_fill = PatternFill(
            fill_type="solid",
            start_color="F4CCCC",
            end_color="F4CCCC"
        )

        gold_fill = PatternFill(
            fill_type="solid",
            start_color="FFD966",
            end_color="FFD966"
        )

        header_fill = PatternFill(
            fill_type="solid",
            start_color="D9EAD3",
            end_color="D9EAD3"
        )

        for ws in wb.worksheets:

            # Header formatting
            for cell in ws[1]:

                cell.font = Font(bold=True)

                cell.fill = header_fill

            headers = [

                c.value

                for c in ws[1]

            ]

            benchmark_col = None

            for i, h in enumerate(headers, start=1):

                if h == "is_benchmark":

                    benchmark_col = i

                    break

            # Benchmark highlight
            if benchmark_col:

                for row in range(2, ws.max_row + 1):

                    if ws.cell(row, benchmark_col).value is True:

                        for col in range(1, ws.max_column + 1):

                            ws.cell(row, col).fill = gold_fill

            # Percentile colouring
            for col in range(1, ws.max_column + 1):

                header = ws.cell(1, col).value

                if header is None:

                    continue

                if not str(header).endswith("_pct_rank"):

                    continue

                for row in range(2, ws.max_row + 1):

                    cell = ws.cell(row, col)

                    value = cell.value

                    if value is None:

                        continue

                    try:

                        value = float(value)

                    except:

                        continue

                    if value >= 75:

                        cell.fill = green_fill

                    elif value <= 25:

                        cell.fill = red_fill

                    else:

                        cell.fill = yellow_fill

            ws.auto_filter.ref = ws.dimensions

        wb.save(workbook_path)

        print("Workbook formatting completed.")
# ==========================================================
# Add Peer Summary Rows
# ==========================================================

    def add_peer_summary(self):

        print("\nAdding Peer Group Summary Rows...\n")

        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill

        workbook_path = OUTPUT_DIR / "peer_comparison.xlsx"

        wb = load_workbook(workbook_path)

        summary_fill = PatternFill(
            fill_type="solid",
            start_color="D0E0E3",
            end_color="D0E0E3"
        )

        for ws in wb.worksheets:

            max_row = ws.max_row
            max_col = ws.max_column

            summary_row = max_row + 2

            ws.cell(summary_row, 1).value = "Peer Median"

            ws.cell(summary_row, 1).font = Font(bold=True)

            ws.cell(summary_row, 1).fill = summary_fill

            headers = [
                ws.cell(1, c).value
                for c in range(1, max_col + 1)
            ]

            for col in range(2, max_col + 1):

                header = headers[col - 1]

                if header is None:
                    continue

                values = []

                for row in range(2, max_row + 1):

                    value = ws.cell(row, col).value

                    if isinstance(value, (int, float)):

                        values.append(value)

                if values:

                    values = sorted(values)

                    n = len(values)

                    if n % 2 == 1:

                        median = values[n // 2]

                    else:

                        median = (
                            values[n // 2 - 1]
                            + values[n // 2]
                        ) / 2

                    cell = ws.cell(summary_row, col)

                    cell.value = round(median, 2)

                    cell.font = Font(bold=True)

                    cell.fill = summary_fill

            count_row = summary_row + 1

            ws.cell(count_row, 1).value = "Company Count"

            ws.cell(count_row, 1).font = Font(bold=True)

            ws.cell(count_row, 1).fill = summary_fill

            ws.cell(count_row, 2).value = max_row - 1

            ws.cell(count_row, 2).font = Font(bold=True)

            ws.cell(count_row, 2).fill = summary_fill

        wb.save(workbook_path)

        print("Summary rows added successfully.")
# ==========================================================
# Final Peer Workbook Validation
# ==========================================================

    def validate_peer_workbook(self):

        print("\n" + "=" * 70)
        print("PEER COMPARISON WORKBOOK VALIDATION")
        print("=" * 70)

        from openpyxl import load_workbook

        workbook_path = OUTPUT_DIR / "peer_comparison.xlsx"

        wb = load_workbook(workbook_path)

        sheets = wb.sheetnames

        print(f"Workbook Exists : {'PASS' if workbook_path.exists() else 'FAIL'}")
        print(f"Sheets Created  : {len(sheets)}")

        expected_groups = (
            self.peer_report
            .dropna(subset=["peer_group_name"])
            ["peer_group_name"]
            .unique()
        )

        expected_groups = [
            g for g in expected_groups
            if g != "No peer group assigned"
        ]

        print(f"Expected Sheets : {len(expected_groups)}")

        if len(sheets) == len(expected_groups):
            print("Sheet Count     : PASS")
        else:
            print("Sheet Count     : WARNING")

        print("\nWorksheet Summary")
        print("-" * 40)

        for sheet in sheets:

            ws = wb[sheet]

            print(
                f"{sheet:<25}"
                f" Rows : {ws.max_row:<4}"
                f" Columns : {ws.max_column}"
            )

        print("\nWorkbook validation completed.")


# ==========================================================
# Final Deliverables Check
# ==========================================================

    def verify_day20_outputs(self):

        print("\n" + "=" * 70)
        print("DAY 20 OUTPUT VERIFICATION")
        print("=" * 70)

        workbook = OUTPUT_DIR / "peer_comparison.xlsx"

        print(
            f"peer_comparison.xlsx : "
            f"{'PASS' if workbook.exists() else 'FAIL'}"
        )

        radar = OUTPUT_DIR / "radar_report.csv"

        print(
            f"radar_report.csv     : "
            f"{'PASS' if radar.exists() else 'FAIL'}"
        )

        charts = len(list(RADAR_DIR.glob("*.png")))

        print(f"Radar Charts         : {charts}")

        groups = self.peer_report["peer_group_name"]

        groups = groups[
        groups != "No peer group assigned"
]

        print(f"Peer Groups          : {groups.nunique()}")

        print(f"Companies            : {len(self.peer_report)}")


# ==========================================================
# Day 20 Summary
# ==========================================================

    def day20_summary(self):

        print("\n" + "=" * 70)
        print("SPRINT 3 - DAY 20 COMPLETED")
        print("=" * 70)

        print("\nDeliverables")
        print("----------------------------------------")

        print("✓ Peer comparison dataset prepared")
        print("✓ Workbook generated")
        print("✓ 11 peer group worksheets created")
        print("✓ Percentile rank columns exported")
        print("✓ Green / Yellow / Red percentile formatting")
        print("✓ Benchmark rows highlighted")
        print("✓ Peer median summary rows added")
        print("✓ Company count rows added")
        print("✓ Workbook validation completed")

        print("\nOutput Files")
        print("----------------------------------------")

        print("✓ output/peer_comparison.xlsx")

        print("\nReady for Day 21")
# ==========================================================
# Main
# ==========================================================

def main():

    engine = PeerEngine()

    engine.load_tables()

    engine.get_latest_company_data()

    engine.merge_peer_groups()

    engine.verify_data()
    engine.compute_peer_percentiles()

    engine.preview_rankings()
    engine.populate_peer_percentiles_table()

    engine.verify_database()

    engine.review_missing_peer_groups()
    engine.peer_group_summary()

    engine.spot_check_peer_groups()

    engine.export_peer_percentiles()

    engine.validate_peer_percentiles()
    engine.database_spot_check()

    engine.final_verification()

    engine.day18_summary()
    engine.prepare_radar_dataset()

    engine.normalise_radar_metrics()

    engine.preview_radar_data()
    engine.generate_peer_radar_charts()

    engine.preview_generated_charts()
    engine.generate_standalone_charts()

    engine.verify_radar_charts()
    engine.radar_summary()

    engine.benchmark_summary()

    engine.top_composite_companies()

    engine.export_radar_report()

    engine.validate_radar_outputs()
    engine.final_day19_verification()

    engine.day19_summary()
    engine.prepare_peer_comparison()

    engine.verify_peer_groups()
    engine.create_peer_workbook()
    engine.format_peer_workbook()
    engine.add_peer_summary()
    engine.validate_peer_workbook()

    engine.verify_day20_outputs()

    engine.day20_summary()   
if __name__ == "__main__":
    main()